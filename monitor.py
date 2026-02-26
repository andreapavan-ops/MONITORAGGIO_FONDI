"""
monitor.py - Script principale di monitoraggio fondi
=====================================================
Orchestrazione del sistema:
1. Legge file Excel con lista fondi
2. Recupera NAV per ogni fondo
3. Calcola indicatori tecnici
4. Genera segnali
5. Invia alert
6. Aggiorna Excel e dashboard
"""

import os
import sys
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
import time
from decimal import Decimal
from pathlib import Path

# Carica variabili .env se presenti (utile quando monitor.py viene eseguito direttamente)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Import moduli locali
from data_fetcher import FundDataFetcher
from technical_analysis import TechnicalAnalyzer
from alerts import AlertSystem
from database import PriceDatabase


# Log globale degli errori consultabile via /api/monitor-log
monitor_log = []


def add_log(message: str):
    """Aggiunge un messaggio al log globale"""
    entry = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
    monitor_log.append(entry)
    # Tieni solo gli ultimi 200 messaggi
    if len(monitor_log) > 200:
        monitor_log.pop(0)
    print(entry)


class FundMonitor:
    """Sistema principale di monitoraggio fondi"""

    def __init__(self, excel_path: str = 'fondi_monitoraggio.xlsx'):
        """
        Inizializza il monitor
        
        Args:
            excel_path: Percorso al file Excel master
        """
        self.excel_path = excel_path
        self.data_fetcher = FundDataFetcher()
        self.analyzer = TechnicalAnalyzer()
        self.alert_system = AlertSystem()

        # Database PostgreSQL per storico prezzi (persistente su Railway)
        self.db = PriceDatabase()

        # Fallback: storage locale per dati storici (se DB non disponibile)
        self.history_path = Path('data/history')
        self.history_path.mkdir(parents=True, exist_ok=True)

        # Carica configurazione
        self.config = self._load_config()
    
    def _load_config(self) -> dict:
        """Carica configurazione dal file Excel"""
        try:
            df_config = pd.read_excel(self.excel_path, sheet_name='CONFIG', header=None)
            config = {}
            for _, row in df_config.iterrows():
                if pd.notna(row[0]) and pd.notna(row[1]):
                    key = str(row[0]).strip()
                    value = row[1]
                    config[key] = value
            return config
        except Exception as e:
            print(f"⚠️ Errore caricamento config: {e}")
            return {
                'Soglia RSI Ipervenduto': 30,
                'Soglia RSI Ipercomprato': 70,
                'Giorni Media Mobile': 15,
                'Min Indicatori Concordi L3': 2
            }
    
    def load_funds(self) -> pd.DataFrame:
        """
        Carica lista fondi dal file Excel

        Returns:
            DataFrame con tutti i fondi (ISIN validi, senza duplicati)
        """
        try:
            df = pd.read_excel(self.excel_path, sheet_name='Fondi')
            total_raw = len(df)

            # Rimuovi righe senza ISIN
            no_isin = df['ISIN'].isna()
            if no_isin.any():
                add_log(f"  ATTENZIONE: {no_isin.sum()} fondi senza ISIN ignorati: {df.loc[no_isin, 'Nome Fondo'].tolist()}")
                df = df[~no_isin]

            # Rimuovi ISIN duplicati (tieni prima occorrenza, logga i rimossi)
            dup_mask = df.duplicated('ISIN', keep='first')
            if dup_mask.any():
                dup_isins = df.loc[dup_mask, 'ISIN'].tolist()
                add_log(f"  ATTENZIONE: {dup_mask.sum()} righe duplicate rimosse (ISIN ripetuti): {dup_isins}")
                df = df[~dup_mask]

            add_log(f"📂 Caricati {len(df)} fondi validi (su {total_raw} righe nel file Excel)")
            return df.reset_index(drop=True)
        except Exception as e:
            print(f"❌ Errore caricamento fondi: {e}")
            return pd.DataFrame()
    
    def get_fund_history(self, isin: str) -> pd.Series:
        """
        Recupera storico prezzi per un fondo dal database PostgreSQL

        Args:
            isin: Codice ISIN del fondo

        Returns:
            Serie pandas con prezzi storici
        """
        # Recupera prezzo odierno (o la data fornita dalla fonte) e salvalo solo se è più recente
        nav_data = self.data_fetcher.get_nav(isin)
        if nav_data and nav_data.get('price') is not None:
            # Fonte può riportare la data del prezzo; altrimenti consideriamo oggi
            fetched_date_str = nav_data.get('date')
            try:
                fetched_date = datetime.fromisoformat(fetched_date_str).date() if fetched_date_str else datetime.now().date()
            except Exception:
                fetched_date = datetime.now().date()

            source = nav_data.get('source', 'FT Markets')

            # Controlla ultima data salvata nel DB (se disponibile)
            try:
                last_date_str = self.db.get_last_price_date(isin)
            except Exception:
                last_date_str = None

            last_date = None
            if last_date_str:
                try:
                    last_date = datetime.fromisoformat(last_date_str).date()
                except Exception:
                    last_date = None

            # Se la data recuperata non è più recente, saltiamo il salvataggio
            if last_date and fetched_date <= last_date:
                print(f"  ℹ️ Dato {fetched_date} non più recente di ultimo salvataggio {last_date}, skip save")
            else:
                # Salva nel database PostgreSQL; se fallisce, salva localmente in data/history
                try:
                    saved = self.db.save_price(isin, fetched_date.strftime('%Y-%m-%d'), nav_data['price'], source)
                except Exception:
                    saved = False

                if not saved:
                    # Fallback: append al file JSON locale per preservare lo storico
                    try:
                        history_file = self.history_path / f"{isin}.json"
                        history = []
                        if history_file.exists():
                            with open(history_file, 'r') as fh:
                                try:
                                    history = json.load(fh)
                                except Exception:
                                    history = []

                        history.append({'date': fetched_date.strftime('%Y-%m-%d'), 'price': nav_data['price'], 'source': source})
                        # Mantieni ultimi 365 giorni
                        if len(history) > 365:
                            history = history[-365:]
                        with open(history_file, 'w') as fh:
                            json.dump(history, fh)
                        print(f"  💾 Prezzo salvato in file locale: {isin} = {nav_data['price']} ({fetched_date.strftime('%Y-%m-%d')})")
                    except Exception as e:
                        print(f"❌ Errore salvataggio locale prezzo {isin}: {e}")

        # Recupera storico dal database (ultimi 100 giorni)
        prices = self.db.get_price_series(isin, days=100)

        if not prices.empty:
            return prices

        # Fallback: prova file JSON locale (per retrocompatibilità)
        history_file = self.history_path / f"{isin}.json"
        if history_file.exists():
            try:
                with open(history_file, 'r') as f:
                    history = json.load(f)
                if history:
                    df = pd.DataFrame(history)
                    return pd.Series(df['price'].values, index=pd.to_datetime(df['date']))
            except Exception:
                pass

        return pd.Series(dtype=float)
    
    def analyze_fund(self, row: pd.Series) -> dict:
        """
        Analizza un singolo fondo
        
        Args:
            row: Riga del DataFrame con dati fondo
        
        Returns:
            Dizionario con risultati analisi
        """
        isin = row['ISIN']
        level = int(row['Livello'])
        
        print(f"  📊 Analisi {row['Nome Fondo'][:40]}...")
        
        # Recupera storico dai file JSON locali (solo dati reali)
        prices = self.get_fund_history(isin)

        if len(prices) < 5:
            # Storico insufficiente - prova Yahoo Finance come ultima risorsa
            df_hist = self.data_fetcher.get_historical_nav(isin, days=30)
            if not df_hist.empty:
                prices = pd.Series(df_hist['nav'].values)
            # Se ancora insufficiente, l'analisi restituirà "dati insufficienti"
        
        # Rileva tipo di asset dalla categoria e crea analyzer dedicato
        categoria = row['Categoria']
        asset_type = TechnicalAnalyzer.detect_asset_type(categoria)
        analyzer = TechnicalAnalyzer(asset_type=asset_type)

        # Esegui analisi con parametri calibrati per il tipo di fondo
        analysis = analyzer.analyze_fund(prices, level=level)
        analysis['asset_type'] = asset_type

        return {
            'isin': isin,
            'nome': row['Nome Fondo'],
            'casa': row['Casa Gestione'],
            'categoria': categoria,
            'livello': level,
            'analysis': analysis
        }
    
    def update_excel(self, results: list):
        """
        Aggiorna il file Excel con i risultati dell'analisi
        Include aggiornamento automatico del livello

        Args:
            results: Lista di risultati analisi
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb['Fondi']

            # Mappa colonne Excel:
            # A(1)=Livello, B(2)=ISIN, C(3)=Nome, D(4)=Casa, E(5)=Categoria,
            # F(6)=Valuta, G(7)=Prezzo, H(8)=MM20, I(9)=RSI, J(10)=Segnale, K(11)=Ultima Modifica
            COL_LIVELLO = 1
            COL_ISIN = 2
            COL_PREZZO = 7
            COL_MM = 8
            COL_RSI = 9
            COL_SEGNALE = 10
            COL_ULTIMA_MODIFICA = 11

            # Mappa ISIN -> riga
            isin_to_row = {}
            for row in range(2, ws.max_row + 1):
                isin = ws.cell(row=row, column=COL_ISIN).value
                if isin:
                    isin_to_row[isin] = row

            level_changes = []

            # Aggiorna dati
            for result in results:
                isin = result['isin']
                analysis = result['analysis']

                if isin in isin_to_row:
                    row = isin_to_row[isin]

                    # Livello (colonna A) - AGGIORNAMENTO AUTOMATICO
                    current_level = result['livello']
                    suggested_level = analysis.get('suggested_level', current_level)
                    level_reason = analysis.get('level_reason', '')

                    if suggested_level != current_level:
                        ws.cell(row=row, column=COL_LIVELLO, value=suggested_level)
                        level_changes.append({
                            'nome': result['nome'],
                            'isin': isin,
                            'from': current_level,
                            'to': suggested_level,
                            'reason': level_reason
                        })
                        # Colora cella livello in base al cambio
                        level_cell = ws.cell(row=row, column=COL_LIVELLO)
                        if suggested_level < current_level:  # Upgrade (es. 3→2 o 2→1)
                            level_cell.fill = PatternFill("solid", fgColor="00B050")
                            level_cell.font = Font(bold=True, color="FFFFFF")
                        else:  # Downgrade (es. 1→2 o 2→3)
                            level_cell.fill = PatternFill("solid", fgColor="FF6600")
                            level_cell.font = Font(bold=True, color="FFFFFF")

                    # Prezzo (colonna G)
                    ws.cell(row=row, column=COL_PREZZO, value=analysis.get('current_price'))

                    # MM20 (colonna H)
                    ws.cell(row=row, column=COL_MM, value=analysis.get('ma'))

                    # RSI (colonna I)
                    ws.cell(row=row, column=COL_RSI, value=analysis.get('rsi'))

                    # Segnale (colonna J)
                    signal = analysis.get('final_signal', 'HOLD')
                    signal_cell = ws.cell(row=row, column=COL_SEGNALE, value=signal)

                    # Colora cella segnale
                    if signal == 'BUY':
                        signal_cell.fill = PatternFill("solid", fgColor="00B050")
                        signal_cell.font = Font(bold=True, color="FFFFFF")
                    elif signal == 'SELL':
                        signal_cell.fill = PatternFill("solid", fgColor="FF0000")
                        signal_cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        signal_cell.fill = PatternFill("solid", fgColor="FFC000")
                        signal_cell.font = Font(bold=True)

                    # Ultima Modifica (colonna K)
                    ws.cell(row=row, column=COL_ULTIMA_MODIFICA, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

            wb.save(self.excel_path)
            print(f"✅ File Excel aggiornato")

            # Log cambi di livello
            if level_changes:
                print(f"\n📊 CAMBI DI LIVELLO AUTOMATICI:")
                for change in level_changes:
                    arrow = "⬆️" if change['to'] < change['from'] else "⬇️"
                    print(f"  {arrow} {change['nome'][:40]}: L{change['from']} → L{change['to']}")
                    print(f"     Motivo: {change['reason']}")

            return level_changes

        except Exception as e:
            print(f"❌ Errore aggiornamento Excel: {e}")
            return []
    
    def generate_dashboard_data(self, results: list) -> dict:
        """
        Genera dati per la dashboard HTML
        
        Args:
            results: Lista risultati analisi
        
        Returns:
            Dizionario con dati dashboard
        """
        dashboard_data = {
            'last_update': datetime.now().isoformat(),
            'summary': {
                'total_funds': len(results),
                'buy_signals': 0,
                'sell_signals': 0,
                'hold_signals': 0
            },
            'levels': {
                1: [],
                2: [],
                3: []
            },
            'categories': {}
        }
        
        for r in results:
            signal = r['analysis'].get('final_signal', 'HOLD')
            # Usa il livello aggiornato (suggested_level) se è cambiato,
            # così i fondi promossi/declassati appaiono nel bucket corretto
            level = r['analysis'].get('suggested_level', r['livello'])
            category = r['categoria']
            
            # Conteggio segnali
            if signal == 'BUY':
                dashboard_data['summary']['buy_signals'] += 1
            elif signal == 'SELL':
                dashboard_data['summary']['sell_signals'] += 1
            else:
                dashboard_data['summary']['hold_signals'] += 1
            
            # Recupera prezzo di ieri dal database
            price_today = r['analysis'].get('current_price')
            price_yesterday = self.db.get_yesterday_price(r['isin'])

            # Calcola variazione percentuale (converti Decimal in float)
            change_pct = None
            if price_today and price_yesterday:
                change_pct = round(((float(price_today) - float(price_yesterday)) / float(price_yesterday)) * 100, 2)

            # Dati per livello
            fund_data = {
                'isin': r['isin'],
                'nome': r['nome'],
                'casa': r['casa'],
                'categoria': category,
                'price': float(price_today) if price_today else None,
                'price_yesterday': float(price_yesterday) if price_yesterday else None,
                'change_pct': change_pct,
                'ma': float(r['analysis'].get('ma')) if r['analysis'].get('ma') is not None else None,
                'rsi': float(r['analysis'].get('rsi')) if r['analysis'].get('rsi') is not None else None,
                'pct_1d': float(r['analysis'].get('pct_change_1d')) if r['analysis'].get('pct_change_1d') is not None else None,
                'pct_1w': float(r['analysis'].get('pct_change_1w')) if r['analysis'].get('pct_change_1w') is not None else None,
                'pct_1m': float(r['analysis'].get('pct_change_1m')) if r['analysis'].get('pct_change_1m') is not None else None,
                'buy_count': int(r['analysis'].get('buy_count', 0)),
                'asset_type': r['analysis'].get('asset_type', 'equity'),
                'conditions': {
                    'trend_ok': bool(r['analysis'].get('level_conditions', {}).get('trend_ok', False)),
                    'rsi_optimal': bool(r['analysis'].get('level_conditions', {}).get('rsi_optimal', False)),
                    'nav_above_bb': bool(r['analysis'].get('level_conditions', {}).get('nav_above_upper_bb', False)),
                    'nav_rising': bool(r['analysis'].get('level_conditions', {}).get('nav_rising', False)),
                    'nav_rising_original': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_original', False)),
                    'nav_rising_alt': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_alt', False)),
                }
            }
            dashboard_data['levels'][level].append(fund_data)

            # Dati per categoria
            if category not in dashboard_data['categories']:
                dashboard_data['categories'][category] = []
            dashboard_data['categories'][category].append(fund_data)
        
        # Salva JSON per dashboard (converte Decimal e numpy types)
        class SafeEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                # Gestione tipi numpy
                import numpy as np
                if isinstance(obj, (np.integer,)):
                    return int(obj)
                if isinstance(obj, (np.floating,)):
                    return float(obj)
                if isinstance(obj, (np.bool_,)):
                    return bool(obj)
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                return super().default(obj)

        # Includi report di salute se disponibile
        if hasattr(self, '_health_report'):
            dashboard_data['health'] = self._health_report

        with open('data/dashboard_data.json', 'w') as f:
            json.dump(dashboard_data, f, indent=2, cls=SafeEncoder)

        return dashboard_data
    
    def _compute_gap_analysis(self, result: dict) -> dict:
        """
        Calcola l'analisi quantitativa del gap tra un fondo L2 e i criteri L1 Pro.

        Per ciascuna delle 4 condizioni L1 Pro restituisce:
        - se è soddisfatta e il dettaglio
        - se non è soddisfatta: il gap numerico e la previsione

        Returns:
            Dizionario con buy_count, conditions (lista), e valori tecnici chiave
        """
        analysis = result['analysis']
        asset_type = analysis.get('asset_type', 'equity')
        profile = TechnicalAnalyzer.PROFILES.get(asset_type, TechnicalAnalyzer.PROFILES['equity'])

        lc = analysis.get('level_conditions', {})
        conditions = []

        # --- Condizione 1: TREND ---
        trend_ok = lc.get('trend_ok', False)
        days_above = lc.get('days_above_ma', 0)
        slope = float(lc.get('ma_slope', 0) or 0)
        distance = float(lc.get('distance_from_ma', 0) or 0)
        max_dist = profile['max_distance_from_ma']

        if trend_ok:
            conditions.append({
                'name': 'TREND', 'ok': True,
                'detail': f"Prezzo sopra MM da {days_above}gg, slope +{slope:.2f}%, distanza {distance:.1f}%",
                'gap_text': None, 'forecast': None
            })
        else:
            gaps, forecasts = [], []
            if not lc.get('price_above_ma_3days'):
                dn = max(0, 3 - days_above)
                gaps.append(f"Sopra MM da {days_above}/3 giorni richiesti")
                forecasts.append(f"Mancano {dn} seduta{'e' if dn != 1 else 'a'} consecutive sopra MM")
            if not lc.get('slope_positive'):
                gaps.append(f"Pendenza MM: {slope:.3f}% (serve > 0)")
                forecasts.append("1–2 giorni di prezzi crescenti possono invertire la pendenza MM")
            if not lc.get('distance_ok'):
                overshoot = distance - max_dist
                gaps.append(f"Distanza MM: {distance:.1f}% (max {max_dist:.1f}%, eccesso {overshoot:.1f}%)")
                forecasts.append(f"Il fondo ha corso troppo veloce: serve consolidamento di ~{overshoot:.1f}% o che la MM recuperi")
            conditions.append({
                'name': 'TREND', 'ok': False,
                'detail': None,
                'gap_text': ' · '.join(gaps) if gaps else 'Condizione non soddisfatta',
                'forecast': ' · '.join(forecasts) if forecasts else None
            })

        # --- Condizione 2: MOMENTUM (RSI) ---
        rsi_val = float(lc.get('rsi') or analysis.get('rsi') or 50)
        rsi_low = profile['rsi_optimal_low']
        rsi_high = profile['rsi_optimal_high']
        rsi_ok = lc.get('rsi_optimal', False)

        if rsi_ok:
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': True,
                'detail': f"RSI {rsi_val:.0f} nel range ottimale [{rsi_low}–{rsi_high}]",
                'gap_text': None, 'forecast': None
            })
        elif rsi_val < rsi_low:
            gap = rsi_low - rsi_val
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': False,
                'detail': None,
                'gap_text': f"RSI {rsi_val:.0f} sotto soglia {rsi_low} (mancano {gap:.0f} punti)",
                'forecast': f"Con {'2–3' if gap < 5 else '3–5'} sessioni positive, RSI può raggiungere la zona [{rsi_low}–{rsi_high}]"
            })
        else:
            gap = rsi_val - rsi_high
            conditions.append({
                'name': 'MOMENTUM (RSI)', 'ok': False,
                'detail': None,
                'gap_text': f"RSI {rsi_val:.0f} sopra soglia {rsi_high} (+{gap:.0f} punti, ipercomprato)",
                'forecast': "Probabile fase di consolidamento/correzione prima del rientro nel range ottimale"
            })

        # --- Condizione 3: VOLATILITÀ (Bande di Bollinger) ---
        bb_ok = lc.get('nav_above_upper_bb', False)
        price = float(analysis.get('current_price') or 0)
        bb = analysis.get('bollinger') or {}
        ma_val = float(analysis.get('ma') or 0)
        bb_upper = float(bb.get('upper') or 0)

        if bb_ok:
            conditions.append({
                'name': 'VOLATILITÀ (BB)', 'ok': True,
                'detail': "Prezzo nella metà superiore delle Bande di Bollinger",
                'gap_text': None, 'forecast': None
            })
        elif price and ma_val and bb_upper:
            if asset_type == 'equity':
                midpoint = (ma_val + bb_upper) / 2
                gap_pct = (midpoint - price) / price * 100
                conditions.append({
                    'name': 'VOLATILITÀ (BB)', 'ok': False,
                    'detail': None,
                    'gap_text': f"Prezzo {price:.4f} sotto midpoint BB {midpoint:.4f} (gap: {gap_pct:.1f}%)",
                    'forecast': f"Serve +{gap_pct:.1f}% per entrare nella zona superiore BB: momentum ancora parzialmente sviluppato"
                })
            else:
                gap_pct = (ma_val - price) / price * 100 if price else 0
                conditions.append({
                    'name': 'VOLATILITÀ (BB)', 'ok': False,
                    'detail': None,
                    'gap_text': f"Prezzo {price:.4f} sotto MM {ma_val:.4f} (gap: {gap_pct:.1f}%)",
                    'forecast': f"Il NAV deve superare la MM ({ma_val:.4f}): manca un +{gap_pct:.1f}%"
                })
        else:
            conditions.append({
                'name': 'VOLATILITÀ (BB)', 'ok': False,
                'detail': None,
                'gap_text': "Dati Bollinger non disponibili (storico insufficiente)",
                'forecast': "Attendere accumulo di 20 giorni di storico per il calcolo delle BB"
            })

        # --- Condizione 4: SETUP (NAV in salita) — doppio criterio ---
        rising_days       = lc.get('rising_days', 0)
        nav_rising_orig   = lc.get('nav_rising_original', lc.get('nav_rising', False))
        nav_rising_alt    = lc.get('nav_rising_alt', False)
        pct_vs_5d         = lc.get('pct_vs_5d', 0.0)
        setup_ok          = lc.get('nav_rising', False)  # OR combinato

        sign_a  = '✅' if nav_rising_orig else '❌'
        sign_b  = '✅' if nav_rising_alt  else '❌'
        pct_str = f"{pct_vs_5d:+.2f}%" if pct_vs_5d != 0.0 else "N/D"
        # Formato HTML: ogni criterio su riga separata
        sub_a   = f"{sign_a} <b>A</b>: {rising_days} gg consecutivi in salita (soglia ≥3)"
        sub_b   = f"{sign_b} <b>B</b>: Prezzo vs 5gg fa {pct_str} (soglia >0%)"
        both    = f"{sub_a}<br>{sub_b}"

        if setup_ok:
            which = "A" if nav_rising_orig else "B"
            conditions.append({
                'name': 'SETUP (NAV)', 'ok': True,
                'detail': f"Criterio <b>{which}</b> passato<br>{sub_a}<br>{sub_b}",
                'gap_text': None, 'forecast': None
            })
        else:
            needed  = max(0, 3 - rising_days)
            fc_a    = f"<b>A</b>: mancano {needed} gg positivi consecutivi"
            fc_b    = (f"<b>B</b>: prezzo deve crescere {abs(pct_vs_5d):.2f}% rispetto a 5gg fa"
                       if pct_vs_5d <= 0 else f"<b>B</b>: prezzo già +{pct_vs_5d:.2f}% su 5gg (in recupero)")
            conditions.append({
                'name': 'SETUP (NAV)', 'ok': False,
                'detail': None,
                'gap_text': both,
                'forecast': f"{fc_a}. {fc_b}"
            })

        return {
            'buy_count': int(analysis.get('buy_count', 0)),
            'conditions': conditions,
            'price': price,
            'ma': ma_val,
            'rsi': rsi_val,
            'pct_1d': analysis.get('pct_change_1d'),
            'pct_1w': analysis.get('pct_change_1w'),
            'pct_1m': analysis.get('pct_change_1m'),
            'asset_type': asset_type,
        }

    def send_alerts(self, results: list):
        """
        Invia alert per segnali significativi.

        - SELL su L1/L2: email individuale urgente
        - BUY: unica email digest giornaliera con:
            * Fondi promossi a L1 (tutte 4 condizioni soddisfatte)
            * Top 3 fondi L2 più vicini a L1 con analisi gap e previsione
        """
        promoted_to_l1 = []   # Fondi (qualsiasi livello) con suggested_level == 1
        near_l1 = []          # Fondi L2 non ancora a L1, candidati top 3

        for r in results:
            signal = r['analysis'].get('final_signal', 'HOLD')
            level = r['livello']
            suggested = r['analysis'].get('suggested_level', level)

            fund_info = {
                'isin': r['isin'],
                'nome': r['nome'],
                'casa': r['casa'],
                'categoria': r['categoria'],
                'livello': level,
                'analysis': r['analysis'],
            }

            # SELL su L1: urgente, email individuale (solo L1)
            if level == 1 and signal == 'SELL':
                add_log(f"  🔴 Alert SELL L1: {r['nome'][:40]}")
                self.alert_system.send_sell_alert(fund_info, r['analysis'])

            # BUY: fondo con tutte 4 condizioni ok (già a L1 o appena promosso)
            elif suggested == 1:
                label = f"L{level}→L1" if level != 1 else "L1 confermato"
                add_log(f"  ⬆️ BUY {label} (digest): {r['nome'][:40]}")
                promoted_to_l1.append(fund_info)

            # Fondi L2 non promossi: candidati top 3 per analisi gap
            elif level == 2:
                gap_info = self._compute_gap_analysis(r)
                near_l1.append({**fund_info, 'gap_info': gap_info})

        # Ordina L2 per buy_count decrescente, prendi top 3
        near_l1_sorted = sorted(near_l1, key=lambda x: x['gap_info'].get('buy_count', 0), reverse=True)
        top3 = near_l1_sorted[:3]

        # Invia unica digest BUY
        if promoted_to_l1 or near_l1:
            add_log(f"  📧 Invio digest BUY: {len(promoted_to_l1)} promossi a L1, {len(top3)} top L2")
            self.alert_system.send_buy_digest(promoted_to_l1, top3)
        else:
            add_log("  ℹ️ Nessun fondo L2 da reportare — digest BUY non inviato")
    
    def run(self, send_daily_report: bool = True):
        """
        Esegue ciclo completo di monitoraggio

        Args:
            send_daily_report: Se True, invia report giornaliero
        """
        import traceback
        add_log("="*50)
        add_log(f"FUND MONITOR - Avvio monitoraggio {datetime.now().strftime('%Y-%m-%d %H:%M')}")

        # 1. Carica fondi
        df_funds = self.load_funds()
        if df_funds.empty:
            add_log("ERRORE: Nessun fondo da monitorare (DataFrame vuoto)")
            return

        add_log(f"Caricati {len(df_funds)} fondi dal file Excel")

        # 2. Analizza ogni fondo
        add_log(f"Inizio analisi di {len(df_funds)} fondi...")
        results = []
        errors = []

        for idx, row in df_funds.iterrows():
            try:
                result = self.analyze_fund(row)
                results.append(result)
                add_log(f"  OK {row['ISIN']} - {row['Nome Fondo'][:30]}")
                time.sleep(0.5)  # Rate limiting
            except Exception as e:
                error_detail = traceback.format_exc()
                errors.append({'isin': row['ISIN'], 'error': str(e), 'traceback': error_detail})
                add_log(f"  ERRORE {row['ISIN']}: {e}")
                add_log(f"  TRACEBACK: {error_detail}")

        add_log(f"Analisi completata: {len(results)} OK, {len(errors)} errori")

        # Salva report di salute per la dashboard e le email
        self._health_report = {
            'timestamp': datetime.now().isoformat(),
            'total_funds': len(df_funds),
            'funds_ok': len(results),
            'funds_error': len(errors),
            'errors': [{'isin': e['isin'], 'error': e['error']} for e in errors],
            'db_available': self.db.is_available() if self.db else False,
            'funds_with_price': sum(1 for r in results if r['analysis'].get('current_price') is not None),
            'funds_no_price': sum(1 for r in results if r['analysis'].get('current_price') is None),
        }

        # 3. Aggiorna Excel
        try:
            add_log("Step 3: Aggiornamento file Excel...")
            self.update_excel(results)
            add_log("Step 3: Excel aggiornato OK")
        except Exception as e:
            add_log(f"Step 3 ERRORE Excel: {e}")
            add_log(traceback.format_exc())

        # 4. Genera dati dashboard
        try:
            add_log(f"Step 4: Generazione dashboard con {len(results)} risultati...")
            os.makedirs('data', exist_ok=True)
            dashboard_data = self.generate_dashboard_data(results)
            total = dashboard_data.get('summary', {}).get('total_funds', '?')
            add_log(f"Step 4: Dashboard generata OK - {total} fondi")
        except Exception as e:
            add_log(f"Step 4 ERRORE Dashboard: {e}")
            add_log(traceback.format_exc())
            # Genera dashboard minima per non bloccare tutto
            dashboard_data = {
                'last_update': datetime.now().isoformat(),
                'summary': {'total_funds': len(results), 'buy_signals': 0, 'sell_signals': 0, 'hold_signals': 0},
                'levels': {1: [], 2: [], 3: []},
                'categories': {}
            }
            # Popola con dati minimi
            for r in results:
                try:
                    fund_data = {
                        'isin': r['isin'], 'nome': r['nome'], 'casa': r['casa'],
                        'categoria': r['categoria'], 'price': float(r['analysis'].get('current_price')) if r['analysis'].get('current_price') else None,
                        'price_yesterday': None, 'change_pct': None,
                        'ma': float(r['analysis'].get('ma')) if r['analysis'].get('ma') else None,
                        'rsi': float(r['analysis'].get('rsi')) if r['analysis'].get('rsi') else None,
                        'pct_1d': float(r['analysis'].get('pct_change_1d')) if r['analysis'].get('pct_change_1d') is not None else None,
                        'pct_1w': float(r['analysis'].get('pct_change_1w')) if r['analysis'].get('pct_change_1w') is not None else None,
                        'pct_1m': float(r['analysis'].get('pct_change_1m')) if r['analysis'].get('pct_change_1m') is not None else None,
                        'buy_count': int(r['analysis'].get('buy_count', 0)),
                        'asset_type': r['analysis'].get('asset_type', 'equity'),
                        'conditions': {
                            'trend_ok': bool(r['analysis'].get('level_conditions', {}).get('trend_ok', False)),
                            'rsi_optimal': bool(r['analysis'].get('level_conditions', {}).get('rsi_optimal', False)),
                            'nav_above_bb': bool(r['analysis'].get('level_conditions', {}).get('nav_above_upper_bb', False)),
                            'nav_rising': bool(r['analysis'].get('level_conditions', {}).get('nav_rising', False)),
                            'nav_rising_original': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_original', False)),
                            'nav_rising_alt': bool(r['analysis'].get('level_conditions', {}).get('nav_rising_alt', False)),
                        }
                    }
                    dashboard_data['levels'][r['livello']].append(fund_data)
                except:
                    pass
            with open('data/dashboard_data.json', 'w') as f:
                json.dump(dashboard_data, f, indent=2)
            add_log(f"Step 4: Dashboard fallback salvata con {len(results)} fondi")

        # 5. Invia alert
        try:
            add_log("Step 5: Invio alert...")
            self.send_alerts(results)
            add_log("Step 5: Alert OK")
        except Exception as e:
            add_log(f"Step 5 ERRORE Alert: {e}")

        # 6. Report giornaliero
        if send_daily_report:
            try:
                add_log("Step 6: Invio report giornaliero...")
                summary = {
                    'buy_signals': dashboard_data['summary']['buy_signals'],
                    'sell_signals': dashboard_data['summary']['sell_signals'],
                    'hold_signals': dashboard_data['summary']['hold_signals'],
                    'level_1': dashboard_data['levels'].get(1, []),
                    'level_2': dashboard_data['levels'].get(2, []),
                    'level_3': dashboard_data['levels'].get(3, [])[:10]
                }
                self.alert_system.send_daily_report(summary)
                add_log("Step 6: Report OK")
            except Exception as e:
                add_log(f"Step 6 ERRORE Report: {e}")

        # 7. Health report via email
        if hasattr(self, '_health_report'):
            try:
                add_log("Step 7: Invio health report...")
                self.alert_system.send_health_report(self._health_report)
                add_log("Step 7: Health report OK")
            except Exception as e:
                add_log(f"Step 7 ERRORE Health report: {e}")

        add_log(f"Monitoraggio completato - {datetime.now().strftime('%H:%M')}")
        add_log("="*50)


def main():
    """Entry point"""
    monitor = FundMonitor()
    monitor.run()


if __name__ == "__main__":
    main()
