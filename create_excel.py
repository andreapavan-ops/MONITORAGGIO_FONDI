from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Fondi"

# Tutti i 70 fondi con Livello 3 iniziale
fondi = [
    # AI & TECNOLOGIA
    (3, "LU1548497772", "Allianz Global Artificial Intelligence AT H2-EUR", "Allianz GI", "AI_Tech", "EUR Hedged"),
    (3, "LU1548497699", "Allianz Global Artificial Intelligence AT EUR", "Allianz GI", "AI_Tech", "EUR"),
    (3, "LU2347504379", "M&G (Lux) Global Artificial Intelligence Fund EUR A", "M&G", "AI_Tech", "EUR"),
    (3, "LU0348723411", "Pictet-Robotics-P EUR", "Pictet", "AI_Tech", "EUR"),
    (3, "LU1861132840", "Pictet-Digital-P EUR", "Pictet", "AI_Tech", "EUR"),
    (3, "LU0302296149", "JPMorgan Funds - US Technology A (acc) EUR hedged", "JPMorgan", "AI_Tech", "EUR Hedged"),
    (3, "LU1681045537", "Amundi Funds - Global Disruption A EUR (C)", "Amundi", "AI_Tech", "EUR"),
    (3, "LU0823421689", "Fidelity Funds - Global Technology Fund A-Euro", "Fidelity", "AI_Tech", "EUR"),
    (3, "LU0099574567", "Fidelity Funds - Global Focus Fund A-Euro", "Fidelity", "AI_Tech", "EUR"),
    (3, "LU0056508442", "BGF World Technology Fund A2 EUR", "BlackRock", "AI_Tech", "EUR"),
    
    # METALLI PREZIOSI
    (3, "LU0175576296", "BGF World Gold Fund A2 EUR", "BlackRock", "Gold", "EUR"),
    (3, "LU0055631609", "Schroder ISF Global Gold A Acc EUR", "Schroders", "Gold", "EUR"),
    (3, "LU0147784465", "DWS Invest Gold and Precious Metals Equities LC", "DWS", "Gold", "EUR"),
    (3, "LU0308790152", "Franklin Gold and Precious Metals Fund A EUR-H1", "Franklin", "Gold", "EUR Hedged"),
    (3, "LU0172583626", "JPMorgan Funds - Global Natural Resources A EUR", "JPMorgan", "Gold", "EUR"),
    (3, "LU0823414940", "Ninety One Global Gold Fund A Acc EUR", "Ninety One", "Gold", "EUR"),
    (3, "CH0104851669", "Pictet CH Precious Metals Fund Physical Gold R EUR", "Pictet", "Gold", "EUR"),
    (3, "LU0496786731", "Pictet-Premium Brands-P EUR", "Pictet", "Gold", "EUR"),
    (3, "LU1240329166", "Amundi Funds - CPR Global Gold Mines A EUR (C)", "Amundi", "Gold", "EUR"),
    (3, "LU0259320728", "Invesco Gold & Special Minerals Fund A EUR", "Invesco", "Gold", "EUR"),
    
    # HEALTHCARE
    (3, "LU0047061557", "Fidelity Funds - Global Health Care Fund A-Euro", "Fidelity", "Healthcare", "EUR"),
    (3, "LU0122379950", "BGF World Healthscience Fund A2 EUR", "BlackRock", "Healthcare", "EUR"),
    (3, "LU0188501257", "Pictet-Health-P EUR", "Pictet", "Healthcare", "EUR"),
    (3, "LU0329203144", "JPMorgan Funds - Global Healthcare A (acc) EUR", "JPMorgan", "Healthcare", "EUR"),
    (3, "LU0256846626", "Schroder ISF Global Healthcare A Acc EUR", "Schroders", "Healthcare", "EUR"),
    (3, "LU1683285164", "Candriam Equities L Biotechnology C EUR Cap", "Candriam", "Healthcare", "EUR"),
    (3, "LU1244893696", "Polar Capital Healthcare Discovery Fund EUR", "Polar Capital", "Healthcare", "EUR"),
    (3, "LU0058720904", "Franklin Biotechnology Discovery Fund A EUR", "Franklin", "Healthcare", "EUR"),
    (3, "LU0261946247", "UBS (Lux) EF Health Care EUR P-acc", "UBS", "Healthcare", "EUR"),
    (3, "LU0114720955", "Janus Henderson Global Life Sciences Fund A2 EUR", "Janus Henderson", "Healthcare", "EUR"),
    
    # ENERGY
    (3, "LU0122376428", "BGF World Energy Fund A2 EUR", "BlackRock", "Energy", "EUR"),
    (3, "LU0280435388", "Schroder ISF Global Energy A Acc EUR", "Schroders", "Energy", "EUR"),
    (3, "LU0348926287", "Pictet-Clean Energy Transition-P EUR", "Pictet", "Energy", "EUR"),
    (3, "LU1892830081", "BGF Sustainable Energy Fund A2 EUR", "BlackRock", "Energy", "EUR"),
    (3, "LU0171289902", "JPMorgan Funds - Global Natural Resources A EUR", "JPMorgan", "Energy", "EUR"),
    (3, "LU0557290854", "Fidelity Funds - Global Industrials Fund A-Euro", "Fidelity", "Energy", "EUR"),
    (3, "LU0109392836", "Franklin Natural Resources Fund A EUR", "Franklin", "Energy", "EUR"),
    (3, "LU1861217949", "Amundi Funds - Global Ecology ESG A EUR (C)", "Amundi", "Energy", "EUR"),
    (3, "LU0209158501", "Invesco Energy Transition Fund A EUR Acc", "Invesco", "Energy", "EUR"),
    (3, "LU0323357649", "DWS Invest Global Infrastructure LC", "DWS", "Energy", "EUR"),
    
    # AZIONARI GLOBALI
    (3, "LU0210534227", "Fidelity Funds - World Fund A-Euro", "Fidelity", "Global_Equity", "EUR"),
    (3, "LU0168343191", "JPMorgan Funds - Global Select Equity A EUR", "JPMorgan", "Global_Equity", "EUR"),
    (3, "LU0072462426", "BGF Global Equity Income Fund A2 EUR", "BlackRock", "Global_Equity", "EUR"),
    (3, "LU0243957239", "Schroder ISF QEP Global Quality A Acc EUR", "Schroders", "Global_Equity", "EUR"),
    (3, "LU0256839704", "Pictet-Global Megatrend Selection-P EUR", "Pictet", "Global_Equity", "EUR"),
    (3, "LU0607514717", "MS INVF Global Opportunity A EUR", "Morgan Stanley", "Global_Equity", "EUR"),
    (3, "LU0149534421", "Amundi Funds - Pioneer Global Equity A EUR (C)", "Amundi", "Global_Equity", "EUR"),
    (3, "LU0203975437", "Comgest Growth World EUR Acc", "Comgest", "Global_Equity", "EUR"),
    (3, "LU0109392752", "Franklin Global Growth Fund A EUR", "Franklin", "Global_Equity", "EUR"),
    (3, "LU1670627758", "Capital Group New Perspective Fund (LUX) B EUR", "Capital Group", "Global_Equity", "EUR"),
    
    # OBBLIGAZIONARI EUR
    (3, "IE00B84J9L26", "PIMCO GIS Income Fund E Class EUR Hedged Acc", "PIMCO", "Bond_EUR", "EUR Hedged"),
    (3, "LU1694789451", "DNCA Invest Alpha Bonds A EUR", "DNCA", "Bond_EUR", "EUR"),
    (3, "LU0048580004", "Fidelity Funds - Euro Bond Fund A-Euro", "Fidelity", "Bond_EUR", "EUR"),
    (3, "LU0012119607", "Pictet-EUR Bonds-P", "Pictet", "Bond_EUR", "EUR"),
    (3, "LU0050427557", "Pictet-EUR Corporate Bonds-P", "Pictet", "Bond_EUR", "EUR"),
    (3, "LU0070917599", "BGF Euro Bond Fund A2 EUR", "BlackRock", "Bond_EUR", "EUR"),
    (3, "LU0210529656", "Schroder ISF EURO Corporate Bond A Acc EUR", "Schroders", "Bond_EUR", "EUR"),
    (3, "LU0248045953", "JPMorgan Funds - EU Government Bond A EUR", "JPMorgan", "Bond_EUR", "EUR"),
    (3, "LU1681040538", "Amundi Funds - Euro Corporate Bond A EUR (C)", "Amundi", "Bond_EUR", "EUR"),
    (3, "LU0996181757", "BlueBay Investment Grade Bond Fund R EUR", "BlueBay", "Bond_EUR", "EUR"),
    
    # MONETARI / RISK OFF
    (3, "LU0128494191", "Pictet-Short-Term Money Market EUR-P", "Pictet", "Money_Market", "EUR"),
    (3, "LU0568621618", "JPMorgan Funds - EUR Liquidity VNAV A (acc)", "JPMorgan", "Money_Market", "EUR"),
    (3, "LU1129459035", "Amundi Euro Liquidity Short Term A EUR (C)", "Amundi", "Money_Market", "EUR"),
    (3, "LU0090865873", "Fidelity Funds - Euro Cash Fund A-Euro", "Fidelity", "Money_Market", "EUR"),
    (3, "LU0011850392", "Pictet-EUR Short Mid-Term Bonds-P", "Pictet", "Money_Market", "EUR"),
    (3, "LU0048580855", "Fidelity Funds - Euro Short Term Bond Fund A-Euro", "Fidelity", "Money_Market", "EUR"),
    (3, "LU0088812606", "Schroder ISF EURO Short Term Bond A Acc EUR", "Schroders", "Money_Market", "EUR"),
    (3, "LU0048621717", "BGF Euro Short Duration Bond Fund A2 EUR", "BlackRock", "Money_Market", "EUR"),
    (3, "LU0248046092", "JPMorgan Funds - Euro Money Market A EUR", "JPMorgan", "Money_Market", "EUR"),
    (3, "LU1681041262", "Amundi Funds - Cash EUR A EUR (C)", "Amundi", "Money_Market", "EUR"),
]

# Header
headers = ["Livello", "ISIN", "Nome Fondo", "Casa Gestione", "Categoria", "Valuta", "Prezzo", "MM15", "RSI", "Segnale", "Ultima Modifica"]
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Dati
for row_idx, fund in enumerate(fondi, 2):
    livello, isin, nome, casa, categoria, valuta = fund
    ws.cell(row=row_idx, column=1, value=livello).border = thin_border
    ws.cell(row=row_idx, column=2, value=isin).border = thin_border
    ws.cell(row=row_idx, column=3, value=nome).border = thin_border
    ws.cell(row=row_idx, column=4, value=casa).border = thin_border
    ws.cell(row=row_idx, column=5, value=categoria).border = thin_border
    ws.cell(row=row_idx, column=6, value=valuta).border = thin_border
    ws.cell(row=row_idx, column=7, value="").border = thin_border  # Prezzo
    ws.cell(row=row_idx, column=8, value="").border = thin_border  # MM15
    ws.cell(row=row_idx, column=9, value="").border = thin_border  # RSI
    ws.cell(row=row_idx, column=10, value="").border = thin_border  # Segnale
    ws.cell(row=row_idx, column=11, value="").border = thin_border  # Ultima Modifica
    
    # Colore alternato righe
    if row_idx % 2 == 0:
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="D6DCE4")
    
    # Colore livello
    livello_cell = ws.cell(row=row_idx, column=1)
    livello_cell.alignment = Alignment(horizontal="center")
    if livello == 1:
        livello_cell.fill = PatternFill("solid", fgColor="00B050")
        livello_cell.font = Font(bold=True, color="FFFFFF")
    elif livello == 2:
        livello_cell.fill = PatternFill("solid", fgColor="FFC000")
        livello_cell.font = Font(bold=True)
    else:
        livello_cell.fill = PatternFill("solid", fgColor="4472C4")
        livello_cell.font = Font(bold=True, color="FFFFFF")

# Larghezza colonne
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 8
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 18

# Freeze panes
ws.freeze_panes = 'A2'

# ===================
# FOGLIO ISTRUZIONI
# ===================
ws_help = wb.create_sheet("ISTRUZIONI")

instructions = [
    ("SISTEMA MONITORAGGIO FONDI - ISTRUZIONI", ""),
    ("", ""),
    ("COME USARE IL FILE", ""),
    ("", ""),
    ("▶ AGGIUNGERE UN FONDO:", "Aggiungi una nuova riga nel foglio 'Fondi' con ISIN, nome, casa, categoria, valuta e Livello (1, 2 o 3)"),
    ("▶ RIMUOVERE UN FONDO:", "Cancella la riga corrispondente"),
    ("▶ PROMUOVERE UN FONDO:", "Cambia il valore nella colonna 'Livello' (es. da 3 a 2, o da 2 a 1)"),
    ("▶ RETROCEDERE UN FONDO:", "Cambia il valore nella colonna 'Livello' (es. da 1 a 2, o da 2 a 3)"),
    ("", ""),
    ("SIGNIFICATO LIVELLI", ""),
    ("", ""),
    ("🟢 LIVELLO 1 (Core):", "Fondi in portafoglio - Monitoraggio completo con alert vendita"),
    ("🟡 LIVELLO 2 (Watchlist):", "Fondi candidati acquisto - Alert su setup promettenti"),
    ("🔵 LIVELLO 3 (Universe):", "Universo monitorato - Alert solo se ≥2 indicatori concordano"),
    ("", ""),
    ("COLONNE AUTOMATICHE", ""),
    ("", ""),
    ("Prezzo:", "Ultimo NAV disponibile (aggiornato dal sistema)"),
    ("MM15:", "Media Mobile 15 giorni"),
    ("RSI:", "Relative Strength Index (14 periodi)"),
    ("Segnale:", "🟢 BUY / 🟡 HOLD / 🔴 SELL"),
    ("Ultima Modifica:", "Data ultimo aggiornamento dati"),
    ("", ""),
    ("CATEGORIE DISPONIBILI", ""),
    ("", ""),
    ("AI_Tech:", "Intelligenza Artificiale e Tecnologia"),
    ("Gold:", "Metalli Preziosi e Oro"),
    ("Healthcare:", "Settore Salute e Biotech"),
    ("Energy:", "Energia tradizionale e rinnovabile"),
    ("Global_Equity:", "Azionari Globali"),
    ("Bond_EUR:", "Obbligazionari in EUR"),
    ("Money_Market:", "Monetari e Risk Off"),
    ("", ""),
    ("CASE CON SWITCH GRATUITO FINECO", ""),
    ("", ""),
    ("Pictet, Fidelity, JPMorgan, Schroders, BlackRock, Amundi, Morgan Stanley", ""),
]

for row_idx, (col1, col2) in enumerate(instructions, 1):
    cell1 = ws_help.cell(row=row_idx, column=1, value=col1)
    cell2 = ws_help.cell(row=row_idx, column=2, value=col2)
    
    if "SISTEMA MONITORAGGIO" in col1:
        cell1.font = Font(bold=True, size=16, color="1F4E79")
    elif col1.endswith(":") and col2 == "":
        cell1.font = Font(bold=True, size=12, color="1F4E79")
    elif col1.startswith("▶") or col1.startswith("🟢") or col1.startswith("🟡") or col1.startswith("🔵"):
        cell1.font = Font(bold=True)

ws_help.column_dimensions['A'].width = 40
ws_help.column_dimensions['B'].width = 80

# ===================
# FOGLIO CONFIG
# ===================
ws_config = wb.create_sheet("CONFIG")

config_data = [
    ("CONFIGURAZIONE SISTEMA", ""),
    ("", ""),
    ("Parametro", "Valore"),
    ("Email Alert", "andreapavan67@gmail.com"),
    ("Orario Monitoraggio", "18:00"),
    ("Soglia RSI Ipervenduto", "30"),
    ("Soglia RSI Ipercomprato", "70"),
    ("Giorni Media Mobile", "15"),
    ("Min Indicatori Concordi L3", "2"),
]

for row_idx, (param, value) in enumerate(config_data, 1):
    cell1 = ws_config.cell(row=row_idx, column=1, value=param)
    cell2 = ws_config.cell(row=row_idx, column=2, value=value)
    
    if row_idx == 1:
        cell1.font = Font(bold=True, size=14, color="1F4E79")
    elif row_idx == 3:
        cell1.font = Font(bold=True)
        cell2.font = Font(bold=True)
        cell1.fill = PatternFill("solid", fgColor="D6DCE4")
        cell2.fill = PatternFill("solid", fgColor="D6DCE4")
    elif row_idx > 3:
        cell2.fill = PatternFill("solid", fgColor="FFFF00")

ws_config.column_dimensions['A'].width = 30
ws_config.column_dimensions['B'].width = 35

wb.save('/home/claude/fund_monitor_system/fondi_monitoraggio.xlsx')
print("File Excel master creato con successo!")
